import os
import re
import hashlib
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# import argparse


# Define una función que reemplace los espacios por saltos de línea, excepto los que están entre corchetes
def replace_spaces_except_in_brackets(s):
    in_bracket = False
    new_str_parts = []
    for part in re.split(r"(\[.*?\])", s):
        if in_bracket:
            new_str_parts.append(part)
        else:
            new_str_parts.append(part.replace(" ", "\n"))
        in_bracket = not in_bracket
    return "".join(new_str_parts)


def get_hash_color(value):
    """Genera un color único basado en el hash del valor."""
    hash_obj = hashlib.md5(str(value).encode())
    return hash_obj.hexdigest()[:6]


class Procesado:

    def __init__(self, input_path, output_path):
        self.input_path = input_path
        self.output_path = output_path

    # (Aquí colocarías las definiciones de process_amrfinder, process_mlst, y process_resfinder
    # con la misma estructura que tienen actualmente pero con un ligero cambio en la firma de la función,
    # como se muestra en el siguiente ejemplo para process_amrfinder)

    def process_amrfinder(self):
        INPUT_PATH = self.input_path
        OUTPUT_PATH = self.output_path

        """
        Esta función procesa los ficheros de salida de amrfinder en el directorio INPUT_PATH 
        y unifica los resultados en un único archivo en OUTPUT_PATH.
    
        :param INPUT_PATH: Ruta al directorio de entrada con los ficheros a procesar.
        :param OUTPUT_PATH: Ruta al directorio donde se guardará el archivo de salida.
        """
        # Verifica si el directorio de entrada existe
        if os.path.isdir(INPUT_PATH):

            # Inicializa una lista vacía para almacenar los datos de cada archivo
            PDATA = []

            # Itera sobre todos los archivos en el directorio de entrada
            for filename in os.listdir(INPUT_PATH):

                # Evita leer archivos que comienzan con '.'
                if not filename.startswith("."):

                    # Obtiene el nombre base del archivo
                    basename = os.path.basename(filename)

                    # Obtiene la variable sNAME dividiendo el nombre base por '_amrfinder'
                    sNAME = basename.split("_amrfinder")[0]

                    # Verifica si el nombre del archivo contiene la cadena 'amrfinder'
                    if "_amrfinder" in basename:

                        # Define el camino completo al archivo
                        filepath = os.path.join(INPUT_PATH, filename)

                        # Intenta leer el archivo en un DataFrame de pandas
                        try:
                            table = pd.read_csv(filepath, sep="\t")

                            # Verifica si hay datos en el DataFrame
                            if not table.empty:

                                # Filtra las filas por 'Element type' y agrupa por 'Gene symbol'
                                sVIR = " ".join(
                                    table[table["Type"] == "VIRULENCE"][
                                        "Element symbol"
                                    ]
                                    .astype(str)
                                    .unique()
                                )
                                sAMR = " ".join(
                                    table[table["Type"] == "AMR"]["Element symbol"]
                                    .astype(str)
                                    .unique()
                                )
                                vSCOPE = " ".join(
                                    table[table["Scope"] == "core"]["Element symbol"]
                                    .astype(str)
                                    .unique()
                                )

                                # Crea un nuevo DataFrame con una fila y las columnas especificadas
                                new_df = pd.DataFrame(
                                    {
                                        "Sample": [sNAME],
                                        "AMR": [sAMR],
                                        "VIRULENCE": [sVIR],
                                        "SCOPE_core": [vSCOPE],
                                    }
                                )

                                # Agrega el nuevo DataFrame a la lista PDATA
                                PDATA.append(new_df)

                        except Exception as e:
                            print(f"Error reading file {filepath}: {e}")

            # Concatena todos los DataFrames en PDATA en un solo DataFrame
            amrfinder_df = pd.concat(PDATA)

            # Guarda el DataFrame resultante en un archivo tsv
            # result_df.to_csv(os.path.join(OUTPUT_PATH, 'amrfinder.tsv'), sep='\t', index=False)
            return amrfinder_df

        else:
            print(f"El directorio {INPUT_PATH} no existe.")

    # Puedes llamar a la función de la siguiente manera:
    # process_amrfinder('/ALMEIDA/PROJECTS/epibac/out/amr_mlst', '/ALMEIDA/PROJECTS/epibac/out/report/input')

    def process_mlst(self):
        INPUT_PATH = self.input_path
        OUTPUT_PATH = self.output_path
        """
        Esta función procesa los archivos de salida de MLST en el directorio INPUT_PATH 
        y unifica los resultados en un único archivo en OUTPUT_PATH.

        :param INPUT_PATH: Ruta al directorio de entrada con los archivos a procesar.
        :param OUTPUT_PATH: Ruta al directorio donde se guardará el archivo de salida.
        """
        # Verifica si el directorio de entrada existe
        if os.path.isdir(INPUT_PATH):

            # Inicializa una lista vacía para almacenar los datos de cada archivo
            PDATA = []

            # Itera sobre todos los archivos en el directorio de entrada
            for filename in os.listdir(INPUT_PATH):

                # Evita leer archivos que comienzan con '.' o que no contienen la etiqueta '_mlst'
                if not filename.startswith(".") and "_mlst" in filename:

                    # Define el camino completo al archivo
                    filepath = os.path.join(INPUT_PATH, filename)

                    # Intenta leer el archivo en un DataFrame de pandas
                    try:
                        with open(filepath, "r") as file:
                            for line in file:
                                # Divide la línea por '\t'
                                data = line.strip().split("\t")

                                # Asigna los primeros tres valores a vNAME, vSCHEME y vST
                                vNAME, vSCHEME, vST = data[:3]

                                # Colapsa el resto de los valores en vMLST
                                vMLST = " ".join(data[3:])

                                # Crea un nuevo DataFrame con una fila y las columnas especificadas
                                new_df = pd.DataFrame(
                                    {
                                        "Sample": [vNAME],
                                        "Scheme_mlst": [vSCHEME],
                                        "ST": [vST],
                                        "MLST": [vMLST],
                                    }
                                )

                                # Agrega el nuevo DataFrame a la lista PDATA
                                PDATA.append(new_df)

                    except Exception as e:
                        print(f"Error reading file {filepath}: {e}")

            # Concatena todos los DataFrames en PDATA en un solo DataFrame
            mlst_df = pd.concat(PDATA)

            # Guarda el DataFrame resultante en un archivo tsv
            # result_df.to_csv(os.path.join(OUTPUT_PATH, 'mlst.tsv'), sep='\t', index=False)
            return mlst_df

        else:
            print(f"El directorio {INPUT_PATH} no existe.")

    # Puedes llamar a la función de la siguiente manera:
    # process_mlst('/ALMEIDA/PROJECTS/epibac/out/amr_mlst', '/ALMEIDA/PROJECTS/epibac/out/report/input')

    def process_resfinder(self):
        INPUT_PATH = self.input_path
        OUTPUT_PATH = self.output_path
        """
        Esta función procesa los archivos de salida de RESFINDER en el directorio INPUT_PATH
        y unifica los resultados en un único archivo en OUTPUT_PATH.
        
        :param INPUT_PATH: Ruta al directorio de entrada con los archivos a procesar.
        :param OUTPUT_PATH: Ruta al directorio donde se guardará el archivo de salida.
        """

        resfinder_path = os.path.join(INPUT_PATH, "resfinder")

        if os.path.isdir(resfinder_path):

            PDATA = []

            for sample_dir in os.listdir(resfinder_path):

                if not sample_dir.startswith("."):

                    vNAME = sample_dir
                    sample_path = os.path.join(resfinder_path, sample_dir)

                    resfinder_file = os.path.join(
                        sample_path, "ResFinder_results_tab.txt"
                    )
                    pheno_file = os.path.join(sample_path, "pheno_table.txt")

                    vGENEresfinder = ""
                    vPHENOresfinder = ""

                    if os.path.isfile(resfinder_file):
                        try:
                            resfinder_df = pd.read_csv(resfinder_file, sep="\t")
                            vGENEresfinder = " ".join(
                                resfinder_df["Resistance gene"].dropna().unique()
                            )
                        except Exception as e:
                            print(
                                f"Error reading ResFinder_results_tab.txt file in {sample_path}: {e}"
                            )

                    if os.path.isfile(pheno_file):
                        try:
                            pheno_df = pd.read_csv(
                                pheno_file,
                                sep="\t",
                                header=None,
                                skiprows=17,
                                names=[
                                    "Antimicrobial",
                                    "Class",
                                    "WGS-predicted phenotype",
                                    "Match",
                                    "Genetic background",
                                ],
                            )
                            pheno_df = pheno_df[
                                pheno_df["WGS-predicted phenotype"] == "Resistant"
                            ]

                            grouped = pheno_df.groupby("Class")
                            pheno_list = []

                            for name, group in grouped:
                                antimicrobials = "-".join(
                                    group["Antimicrobial"].unique()
                                )
                                pheno_list.append(f"{antimicrobials}[{name}]")

                            vPHENOresfinder = " ".join(pheno_list)

                        except Exception as e:
                            print(
                                f"Error reading pheno_table.txt file in {sample_path}: {e}"
                            )

                    new_df = pd.DataFrame(
                        {
                            "Sample": [vNAME],
                            "GENE_resfinder": [vGENEresfinder],
                            "PHENO_resfinder": [vPHENOresfinder],
                        }
                    )
                    PDATA.append(new_df)

            resfinder_df = pd.concat(PDATA)
            # result_df.to_csv(os.path.join(OUTPUT_PATH, 'resfinder.tsv'), sep='\t', index=False)
            return resfinder_df

        else:
            print(f"El directorio {resfinder_path} no existe.")

    # Para llamar la función:
    # process_resfinder('/ALMEIDA/PROJECTS/epibac/out/amr_mlst', '/ALMEIDA/PROJECTS/epibac/out/report/input')

    def merge_results(self, mlst_df, amrfinder_df, resfinder_df):
        try:
            # Fusionar DataFrames en el orden mlst, amrfinder, resfinder
            merged_df = mlst_df.merge(amrfinder_df, on="Sample", how="outer").merge(
                resfinder_df, on="Sample", how="outer"
            )

            # Obtener la fecha actual en el formato AAMMDD
            current_date = datetime.now().strftime("%y%m%d")

            # Ordenamos resultados
            merged_df = merged_df.sort_values(
                by=["Sample", "ST"], ascending=[True, True]
            )

            # Escribir el DataFrame combinado en nuevos archivos
            # merged_df.to_csv(os.path.join(self.output_path, 'epibac.tsv'), sep='\t', index=False)
            # merged_df.to_excel(os.path.join(self.output_path, f'{current_date}_EPIBAC.xlsx'), index=False)
            merged_df.to_csv(os.path.join(snakemake.output[1]), sep="\t", index=False)
            # merged_df.to_excel(os.path.join(snakemake.output[2]), index=False)

            # Aplica la función solamente a la columna 'PHENO_resfinder'
            merged_df["PHENO_resfinder"] = merged_df["PHENO_resfinder"].apply(
                replace_spaces_except_in_brackets
            )

            merged_df.to_excel(snakemake.output[2], index=False)

            # Abriendo el archivo nuevamente para ajustar los anchos de las columnas
            book = load_workbook(snakemake.output[2])

            # Obteniendo la hoja activa
            sheet = book.active

            # Ajustando los anchos de las columnas
            col_widths = [18, 22, 10, 50, 45, 30, 40, 40, 60]
            for i, col_width in enumerate(col_widths):
                sheet.column_dimensions[chr(65 + i)].width = col_width

            # Estableciendo el estilo del encabezado
            font = Font(name="Calibri", bold=True)
            alignment = Alignment(wrap_text=True, vertical="top")

            # Crear un diccionario para almacenar los colores únicos para cada valor único
            unique_colors = {}

            # Asignar un color a cada celda en la columna 'ST' basado en su valor
            for row in sheet.iter_rows(min_row=2):
                cell = row[2]  # Número de la columna ST 0 1 2
                if cell.value not in unique_colors:
                    unique_colors[cell.value] = get_hash_color(cell.value)
                cell.fill = PatternFill(
                    start_color=unique_colors[cell.value],
                    end_color=unique_colors[cell.value],
                    fill_type="solid",
                )

            # Ajustando la altura de las filas basándose en el contenido
            for row in sheet.iter_rows(min_row=2):
                max_line_count = 1
                for cell in row:
                    cell.alignment = alignment
                    if cell.value and isinstance(cell.value, str):
                        line_count = cell.value.count("\n") + 1
                        if line_count > max_line_count:
                            max_line_count = line_count
                sheet.row_dimensions[row[0].row].height = (
                    max_line_count * 15
                )  # Ajusta el 15 según sea necesario

            # Poner el encabezado en azul claro
            for cell in sheet[1]:
                cell.fill = PatternFill(
                    start_color="D6E4FF", end_color="D6E4FF", fill_type="solid"
                )

            # Poner la primera columna en gris claro y en negrita
            grey_fill = PatternFill(
                start_color="ededed", end_color="ededed", fill_type="solid"
            )
            bold_font = Font(bold=True)
            for row in sheet.iter_rows(min_row=2):
                cell = row[0]
                cell.fill = grey_fill
                cell.font = bold_font

            # Guardando los cambios
            book.save(snakemake.output[2])

        except Exception as e:
            print(f"Ocurrió un error durante la fusión de archivos: {e}")


if __name__ == "__main__":
    # parser = argparse.ArgumentParser(description="Script para procesar y combinar los resultados de amrfinder, mlst y resfinder")
    # parser.add_argument("input", help="Ruta al directorio de entrada con los archivos a procesar")
    # parser.add_argument("output", help="Ruta al directorio donde se guardarán los archivos de salida")

    with open(snakemake.log[0], "w") as f:
        sys.stderr = sys.stdout = f
        # args = parser.parse_args()
        # procesado = Procesado(args.input, args.output)
        procesado = Procesado(snakemake.params.input, snakemake.output[0])

        mlst_df = procesado.process_mlst()
        amrfinder_df = procesado.process_amrfinder()
        resfinder_df = procesado.process_resfinder()

        procesado.merge_results(mlst_df, amrfinder_df, resfinder_df)
